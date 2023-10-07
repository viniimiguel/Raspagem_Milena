from selenium import webdriver
from time import sleep
from selenium.webdriver.common.by import By
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from selenium.common.exceptions import NoSuchElementException


class Milena():
    def __init__(self):
        self.driver = webdriver.Chrome()
        self.site_link = 'https://www.millenamoveiseeletro.com.br/eletrodomesticos'
        

    def main(self):
        self.abre()
        sleep(2)
        self.mais()
        sleep(2)
        self.raspa()
        sleep(2)
        self.cria_planilhas()
        sleep(2)
        self.envia_email()
        sleep(131231231)

    def abre(self):
        self.driver.get(self.site_link)
        self.driver.maximize_window()
        sleep(2)


    def mais(self):
        passa = {
            'XP':{
                'mais': '#product-search-results > div:nth-child(2) > div.col-sm-12.col-md-9.p > div.row.product-grid.product-tile-plp-space > div.col-12.p-0.grid-footer > div > button'
            }          
        }
        while True:
            try:
                mais = self.driver.find_element(By.CSS_SELECTOR, passa['XP']['mais'])
                mais.click()
                sleep(1)

            except NoSuchElementException:
                try:
                    break
                except:
                    print('algo deu errado')

            except Exception as e:
                print(f'error {e}')

    def raspa(self):
        global armazena_nome, armazena_preco
        contador = 1
        armazena_nome = []
        armazena_preco = []
        while True:
            try:
                produtos = {
                    'XP':{
                        'nome': f'/html/body/div[2]/div/div[2]/div/div/div[1]/div[2]/div[2]/div[2]/div[{contador}]/div/div/div[2]/div[1]/a',        
                        'preco': f'/html/body/div[2]/div/div[2]/div/div/div[1]/div[2]/div[2]/div[2]/div[{contador}]/div/div/div[2]/div[2]/span'
                    }
                }

                nome = self.driver.find_element(By.XPATH, produtos['XP']['nome']).text
                preco = self.driver.find_element(By.XPATH, produtos['XP']['preco']).text

                armazena_nome.append(nome)
                armazena_preco.append(preco)

                print(nome)
                print(preco)
                sleep(1)
                contador += 1
               
                
            except NoSuchElementException:
                print('nao tem mais elementos!')
                break

            except Exception as e:
                print(f'error {e}')

    def cria_planilhas(self):
        planilha = openpyxl.Workbook()
        mili = planilha.active
        mili.title = 'Eletro'
        mili['A1'] = 'Nome'
        mili['B1'] = 'Preco'

        for index, (nome, preco) in enumerate(zip(armazena_nome, armazena_preco), start=2):
            mili.cell(column=1, row=index, value=nome)
            mili.cell(column=2, row=index, value=preco)


        planilha.save('planilha_de_preco.xlsx')
        print('Planilha criada com sucesso!')
    
    def envia_email(self):
        smtp_server = 'smtp.gmail.com'
        smtp_port = 587
        email_sender =input('digite o seu gmail: ')
        email_password =('ufiwahyqoyzgaihn')
        recipient_email = input('pra que voce quer enviar o email?: ') 
        email_subject = 'planilha de dados'


        email_body = 'segue em anaxo a planilha de dados: '


        msg = MIMEMultipart()
        msg['From'] = email_sender
        msg['To'] = recipient_email
        msg['Subject'] = email_subject
        msg.attach(MIMEText(email_body, 'plain'))

        filename = 'planilha_de_preco.xlsx'
        attachment = open(filename, 'rb')

        part = MIMEBase('application', 'octet-stream')
        part.set_payload((attachment).read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', "attachment; filename= %s" % filename)

        msg.attach(part)

        try:

            server = smtplib.SMTP(smtp_server, smtp_port)
            server.starttls()
            
            server.login(email_sender, email_password)
            
            server.sendmail(email_sender, recipient_email, msg.as_string())
            
            server.quit()
            
            print('E-mail com anexo enviado com sucesso!')
        except Exception as e:
            print('Erro ao enviar o e-mail:', str(e))

milena = Milena()
milena.main()
